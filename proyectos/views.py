from django.shortcuts import render,redirect, get_object_or_404
from .forms import ProyectosForm
from django.views.generic import ListView, DetailView
from .models import Proyectos
from contabilidad.models import Ingresos, GastosGenerales, GastosVehiculos, GastosMateriales, GastosManoObra, GastosEquipos, GastosSeguridad
from empleados.models import Salario, Asistencias
from contabilidad.views import recalcular_totales_proyecto
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from decimal import Decimal

from .graficas import grafica_ingresos_vs_gastos_semanales, grafica_ingresos_vs_gastos, grafica_gastos_categoria

import io
from django.http import HttpResponse
from openpyxl import Workbook

# Create your views here.

class ProyectosListView(LoginRequiredMixin,ListView):
    model = Proyectos
    template_name = "proyectos/proyectos.html"
    context_object_name = "projects"

    def get_queryset(self):
        # Obtener la queryset inicial
        queryset = Proyectos.objects.all()

        # Obtener los parámetros de filtro de la URL o de un formulario GET
        estatus = self.request.GET.get('estatus')
        empresa = self.request.GET.get('empresa')
        proyecto = self.request.GET.get('proyecto')
        total_min = self.request.GET.get('total_min')
        total_max = self.request.GET.get('total_max')

        # Aplicar los filtros si los valores no son nulos
        if estatus:
            queryset = queryset.filter(estatus=estatus)
        
        if empresa:
            queryset = queryset.filter(empresa__icontains=empresa)
        
        if proyecto:
            queryset = queryset.filter(proyecto__icontains=proyecto)
        
        if total_min:
            queryset = queryset.filter(total__gte=total_min)
        
        if total_max:
            queryset = queryset.filter(total__lte=total_max)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Calcular la suma del campo 'total' de todos los proyectos filtrados
        total_monto_neto = self.get_queryset().aggregate(total=Coalesce(Sum('total'),Decimal('0.00')))['total']
        
        # Pasar la suma total al contexto
        context['total_monto_neto'] = total_monto_neto
        
        # Pasar los filtros actuales para mantener el estado del formulario
        context['estatus'] = self.request.GET.get('estatus', '')
        context['empresa'] = self.request.GET.get('empresa', '')
        context['proyecto'] = self.request.GET.get('proyecto', '')
        context['total_min'] = self.request.GET.get('total_min', '')
        context['total_max'] = self.request.GET.get('total_max', '')

        return context

    
class ProyectosDetailView(LoginRequiredMixin,DetailView):
    model = Proyectos
    template_name = "proyectos/proyectos_detalles.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proyecto = self.get_object()

        # ----------------------------------------------------------------------#
        # Llamar a la función recalcular_totales_proyecto que está en views.py de contabilidad
        totales = recalcular_totales_proyecto(proyecto.id)
        # Añadir los valores recalculados al contexto
        context.update(totales)
        # ----------------------------------------------------------------------#

        # context['graph_json'] = grafica_ingresos_vs_gastos_semanales(proyecto.id)
        # context['graph_json2'] = grafica_ingresos_vs_gastos(proyecto.id)
        # context['graph_json3'] = grafica_gastos_categoria(proyecto.id)

        return context

@login_required
def registrar_proyecto(request):
    registrar_proyectos_form = ProyectosForm()
    if request.method == "POST":
        registrar_proyectos_form = ProyectosForm(request.POST)
        
        if registrar_proyectos_form.is_valid():
            registrar_proyectos_form.save()
            return redirect('proyectos:proyectos')
        else:
            registrar_proyectos_form = ProyectosForm()

    return render(request,"proyectos/registrar_proyecto.html",{'proyectos_form':registrar_proyectos_form})


@login_required
def toggle_estatus_proyecto(request, slug):
    # Obtener el proyecto por su slug
    proyecto = get_object_or_404(Proyectos, slug=slug)

    # Alternar el estatus entre True (Activo) y False (Inactivo)
    proyecto.estatus = not proyecto.estatus
    proyecto.save()

    # Redirigir a la página de detalle del proyecto o a la lista de proyectos
    return redirect('proyectos:proyectos')

@login_required
def export_proyectos_to_excel(request):
    # Crea un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proyectos'

    # Añade encabezados
    headers = ['ID', 'Nombre', 'Empresa', 'Estatus', 'Resultados']  # Ajusta según tus campos
    ws.append(headers)

    # Añade datos de los proyectos
    proyectos = Proyectos.objects.all()  # Obtén todos los proyectos
    for proyecto in proyectos:
        ws.append([proyecto.id, proyecto.proyecto, proyecto.empresa, proyecto.estatus, proyecto.total])  # Ajusta según tus campos

    # Guarda el libro de trabajo en un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Prepara la respuesta HTTP
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos.xlsx'
    return response


@login_required
def export_project_details_to_excel(request, proyecto_slug):
    # Obtén el proyecto
    proyecto = Proyectos.objects.get(slug=proyecto_slug)
    
    # Crea un libro de trabajo y varias hojas
    wb = Workbook()

    # Hoja de Proyectos
    ws1 = wb.active
    ws1.title = 'Proyectos'
    headers1 = ['ID', 'Proyecto', 'Empresa', 'Estatus', 'Total']
    ws1.append(headers1)
    ws1.append([proyecto.id, proyecto.proyecto, proyecto.empresa, proyecto.estatus, proyecto.total])

    # Hoja de Ingresos
    ws2 = wb.create_sheet(title='Ingresos')
    headers2 = ['ID', 'Concepto', 'Ingreso', 'IVA', 'Referencia', 'Fecha']
    ws2.append(headers2)
    ingresos = Ingresos.objects.filter(proyecto=proyecto)
    for ingreso in ingresos:
        fecha_ingreso = ingreso.fecha
        ws2.append([ingreso.id, ingreso.concepto, ingreso.monto, ingreso.iva, ingreso.referencia, fecha_ingreso])

    # Hoja de Gastos Vehículos
    ws3 = wb.create_sheet(title='Gastos Vehículos')
    headers3 = ['ID', 'Vehículo', 'Cantidad Combustible', 'Monto', 'IVA', 'Ubicación', 'Proveedor', 'Conductor', 'Fecha']
    ws3.append(headers3)
    gastos_vehiculos = GastosVehiculos.objects.filter(proyecto=proyecto)
    for gasto in gastos_vehiculos:
        fecha_gasto = gasto.fecha
        ws3.append([
            gasto.id,
            gasto.vehiculo.vehiculo,  # Extrae el nombre del vehículo
            gasto.cantidad_combustible,
            gasto.monto,
            gasto.iva,
            gasto.ubicacion,
            gasto.proveedor,
            gasto.conductor,
            fecha_gasto
        ])

    # Hoja de Gastos Generales
    ws4 = wb.create_sheet(title='Gastos Generales')
    headers4 = ['ID', 'Concepto', 'Comprador', 'Monto', 'IVA', 'Notas', 'Proveedor', 'Fecha']
    ws4.append(headers4)
    gastos_generales = GastosGenerales.objects.filter(proyecto=proyecto)
    for gasto in gastos_generales:
        fecha_gasto = gasto.fecha
        ws4.append([gasto.id, gasto.concepto, gasto.comprador, gasto.monto, gasto.iva, gasto.notas, gasto.proveedor, fecha_gasto])

    # Hoja de Gastos Materiales
    ws4 = wb.create_sheet(title='Gastos Materiales')
    headers4 = ['ID', 'Concepto', 'Comprador', 'Monto', 'IVA', 'Descripción', 'Proveedor', 'Fecha']
    ws4.append(headers4)
    gastos_materiales = GastosMateriales.objects.filter(proyecto=proyecto)
    for gasto in gastos_materiales:
        fecha_gasto_material = gasto.fecha
        ws4.append([gasto.id, gasto.concepto, gasto.comprador, gasto.monto, gasto.iva, gasto.descripcion, gasto.proveedor, fecha_gasto_material])

    # Hoja de Gastos Seguridad
    ws4 = wb.create_sheet(title='Gastos Seguridad')
    headers4 = ['ID', 'Concepto', 'Comprador', 'Monto', 'IVA', 'Descripción', 'Proveedor', 'Fecha']
    ws4.append(headers4)
    gastos_seguridad = GastosSeguridad.objects.filter(proyecto=proyecto)
    for gasto in gastos_seguridad:
        fecha_gasto_seguridad = gasto.fecha
        ws4.append([gasto.id, gasto.concepto, gasto.comprador, gasto.monto, gasto.iva, gasto.descripcion, gasto.proveedor, fecha_gasto_seguridad])

    # Hoja de Gastos Mano de obra
    ws4 = wb.create_sheet(title='Gastos Mano de Obra')
    headers4 = ['ID', 'Nómina', 'IMSS', 'INFONAVIT', 'ISN', 'ISR', 'Horas extras', 'Monto', 'Lote', 'Fecha']
    ws4.append(headers4)
    gastos_mano_obra = GastosManoObra.objects.filter(proyecto=proyecto)
    for gasto in gastos_mano_obra:
        fecha_gasto_mano_obra = gasto.fecha
        ws4.append([
            gasto.id, 
            gasto.nomina, 
            gasto.imss, 
            gasto.infonavit, 
            gasto.isn, 
            gasto.isr,
            gasto.horas_extras, 
            gasto.monto,
            gasto.lote,
            fecha_gasto_mano_obra])
        
    # Hoja de nóminas. Estas son una extensión de mano de obra
    ws4 = wb.create_sheet(title='Gastos nóminas')
    headers4 = ['ID', 'RFC', 'Nombres', 'Apellido paterno', 'Apellido materno', 'Salario', 'IMSS', 'INFONAVIT', 'ISR', 'Horas extras', 'Lote']
    ws4.append(headers4)
    salarios = Salario.objects.filter(proyecto=proyecto)
    for salario in salarios:
        ws4.append([
            salario.id,
            salario.empleado.rfc,
            salario.empleado.nombres,
            salario.empleado.apellido_paterno,
            salario.empleado.apellido_materno,
            salario.salario, 
            salario.imss, 
            salario.infonavit, 
            salario.isr,
            salario.horas_extras, 
            salario.lote
            ])

    # Hoja de Gastos de equipos
    ws4 = wb.create_sheet(title='Gastos Equipos')
    headers4 = ['ID', 'Concepto', 'Comprador', 'Tiempo de renta', 'Monto', 'IVA', 'Descripción', 'Proveedor', 'Fecha']
    ws4.append(headers4)
    gastos_equipos = GastosEquipos.objects.filter(proyecto=proyecto)
    for gasto in gastos_equipos:
        fecha_gasto_equipos = gasto.fecha
        ws4.append([gasto.id, gasto.concepto, gasto.comprador, gasto.tiempo_renta, gasto.monto, gasto.iva, gasto.descripcion, gasto.proveedor, fecha_gasto_equipos])

    # Hoja de Asistencias
    ws4 = wb.create_sheet(title='Asistencias')
    headers4 = ['ID', 'Nombre', 'Fecha', 'Asistencia', 'Horas Extras']
    ws4.append(headers4)
    asistencias_list = Asistencias.objects.filter(proyecto=proyecto)
    for asistencia in asistencias_list:
        ws4.append([asistencia.id, asistencia.empleado.rfc, asistencia.fecha, asistencia.asistencias, asistencia.horas_extras])

    # Guarda el libro de trabajo en un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Prepara la respuesta HTTP
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={proyecto_slug}_detalles.xlsx'
    return response


# ---------------------------------------------------------------------------------- #
# ---------------------------------------------------------------------------------- #
# ---------------------------------------------------------------------------------- #

def actualizar_progreso(request, slug):
    if request.method == "POST":
        progreso = request.POST.get('progreso')
        
        if progreso:
            proyecto = get_object_or_404(Proyectos, slug=slug)
            proyecto.progreso = progreso
            proyecto.save()

        return redirect('proyectos:detalles_proyecto', slug=proyecto.slug)  

    return render(request, 'proyectos/proyectos_detalles.html')